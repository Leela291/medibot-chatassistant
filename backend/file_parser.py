# backend/file_parser.py
"""
Parse uploaded patient records — PDF, images, CSV/Excel.
Extracts text content and returns a structured summary for the LLM.
"""
import os
import csv
import io
import json
import traceback


def parse_uploaded_file(file_storage) -> dict:
    """
    Parse an uploaded file and extract text content.

    Args:
        file_storage: Flask FileStorage object

    Returns:
        {
            "success":  bool,
            "filename": str,
            "filetype": str,
            "content":  str,   # extracted text
            "error":    str | None,
        }
    """
    filename = file_storage.filename or "unknown"
    ext = os.path.splitext(filename)[1].lower()

    try:
        if ext == ".pdf":
            return _parse_pdf(file_storage, filename)
        elif ext in (".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp", ".heic"):
            return _parse_image(file_storage, filename)
        elif ext == ".csv":
            return _parse_csv(file_storage, filename)
        elif ext in (".xls", ".xlsx"):
            return _parse_excel(file_storage, filename)
        elif ext in (".txt", ".text"):
            return _parse_text(file_storage, filename)
        elif ext == ".json":
            return _parse_json(file_storage, filename)
        else:
            return {
                "success": False,
                "filename": filename,
                "filetype": ext,
                "content": "",
                "error": f"Unsupported file type: {ext}. Supported: PDF, PNG, JPG, CSV, XLS, XLSX, TXT, JSON",
            }
    except Exception as e:
        return {
            "success": False,
            "filename": filename,
            "filetype": ext,
            "content": "",
            "error": f"Error parsing {filename}: {str(e)}",
        }


# ── PDF Parser ───────────────────────────────────────────────────
def _parse_pdf(file_storage, filename):
    """Extract text from PDF using PyPDF2 (pure Python, no system deps)."""
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        return {
            "success": False,
            "filename": filename,
            "filetype": ".pdf",
            "content": "",
            "error": "PyPDF2 not installed. Run: pip install PyPDF2",
        }

    reader = PdfReader(file_storage.stream)
    pages_text = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and text.strip():
            pages_text.append(f"--- Page {i+1} ---\n{text.strip()}")

    if not pages_text:
        return {
            "success": True,
            "filename": filename,
            "filetype": ".pdf",
            "content": "[PDF contained no extractable text — it may be a scanned document. Try uploading as an image instead.]",
            "error": None,
        }

    full_text = "\n\n".join(pages_text)
    # Truncate very long documents
    if len(full_text) > 8000:
        full_text = full_text[:8000] + "\n\n[... Document truncated for processing ...]"

    return {
        "success": True,
        "filename": filename,
        "filetype": ".pdf",
        "content": full_text,
        "error": None,
    }

# ── Image Parser (OCR) ──────────────────────────────────────────
def _parse_image(file_storage, filename):
    """Extract text from images using pytesseract OCR and base64-encode for vision processing."""
    import base64

    # Determine the correct MIME type for Gemini
    ext = os.path.splitext(filename)[1].lower()
    mime_map = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".heic": "image/heic"
    }
    mime_type = mime_map.get(ext, "image/jpeg") # Default to jpeg

    # 1. Base64 encode the image
    try:
        file_storage.stream.seek(0)
        img_bytes = file_storage.stream.read()
        base64_str = base64.b64encode(img_bytes).decode('utf-8')
        file_storage.stream.seek(0)  # Reset stream pointer
    except Exception as e:
        base64_str = None
        print(f"[file_parser Error] Failed to base64-encode image: {e}")

    # Try pytesseract for OCR first (if it's a medical report)
    try:
        from PIL import Image
        import pytesseract

        file_storage.stream.seek(0)
        img = Image.open(file_storage.stream)
        w, h = img.size
        text = pytesseract.image_to_string(img)

        if text and text.strip():
            content = f"[OCR extracted text from image: {filename}]\n\n{text.strip()}"
            if len(content) > 8000:
                content = content[:8000] + "\n\n[... Text truncated ...]"
            return {
                "success": True,
                "filename": filename,
                "filetype": ext,
                "mime_type": mime_type,  # <--- Added dynamically
                "content": content,
                "base64_image": base64_str,
                "is_image": True,
                "error": None,
                "is_document": len(text.strip()) > 100,
                "width": w,
                "height": h,
            }
    except ImportError:
        pass
    except Exception:
        pass

    # Fallback: return visual reference for LLM
    try:
        from PIL import Image

        file_storage.stream.seek(0)
        img = Image.open(file_storage.stream)
        w, h = img.size
        mode = img.mode
        return {
            "success": True,
            "filename": filename,
            "filetype": ext,
            "mime_type": mime_type, # <--- Added dynamically
            "content": f"[Visual Image Uploaded: {filename} (Dimensions: {w}x{h}, Mode: {mode})]",
            "base64_image": base64_str,
            "is_image": True,
            "error": None,
            "width": w,
            "height": h,
        }
    except ImportError:
        return {
            "success": True,
            "filename": filename,
            "filetype": ext,
            "mime_type": mime_type, # <--- Added dynamically
            "content": f"[Visual Image Uploaded: {filename}]",
            "base64_image": base64_str,
            "is_image": True,
            "error": None,
            "width": w,
            "height": h,
        }

# ── CSV Parser ───────────────────────────────────────────────────
def _parse_csv(file_storage, filename):
    """Parse CSV file and return formatted table content."""
    raw = file_storage.stream.read()
    # Try UTF-8 first, then latin-1
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        return {
            "success": False,
            "filename": filename,
            "filetype": ".csv",
            "content": "",
            "error": "Could not decode CSV file.",
        }

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    if not rows:
        return {
            "success": True,
            "filename": filename,
            "filetype": ".csv",
            "content": "[Empty CSV file]",
            "error": None,
        }

    # Format as readable table
    header = rows[0]
    data_rows = rows[1:]

    lines = [f"[CSV File: {filename}]"]
    lines.append(f"Columns: {', '.join(header)}")
    lines.append(f"Total rows: {len(data_rows)}")
    lines.append("")

    # Show first 50 rows max
    display_rows = data_rows[:50]
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(["---"] * len(header)) + " |")
    for row in display_rows:
        # Pad row to match header length
        padded = row + [""] * (len(header) - len(row))
        lines.append("| " + " | ".join(padded[:len(header)]) + " |")

    if len(data_rows) > 50:
        lines.append(f"\n[... Showing 50 of {len(data_rows)} rows ...]")

    content = "\n".join(lines)
    if len(content) > 8000:
        content = content[:8000] + "\n\n[... Data truncated ...]"

    return {
        "success": True,
        "filename": filename,
        "filetype": ".csv",
        "content": content,
        "error": None,
    }


# ── Excel Parser ─────────────────────────────────────────────────
def _parse_excel(file_storage, filename):
    """Parse Excel files using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        return {
            "success": False,
            "filename": filename,
            "filetype": os.path.splitext(filename)[1].lower(),
            "content": "",
            "error": "openpyxl not installed. Run: pip install openpyxl",
        }

    wb = openpyxl.load_workbook(file_storage.stream, read_only=True, data_only=True)
    lines = [f"[Excel File: {filename}]"]
    lines.append(f"Sheets: {', '.join(wb.sheetnames)}")
    lines.append("")

    for sheet_name in wb.sheetnames[:5]:  # Limit to 5 sheets
        ws = wb[sheet_name]
        lines.append(f"=== Sheet: {sheet_name} ===")

        rows_data = []
        for row in ws.iter_rows(max_row=51, values_only=True):
            rows_data.append([str(cell) if cell is not None else "" for cell in row])

        if not rows_data:
            lines.append("[Empty sheet]")
            continue

        header = rows_data[0]
        lines.append(f"Columns: {', '.join(header)}")
        lines.append("")
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")

        for row in rows_data[1:]:
            padded = row + [""] * (len(header) - len(row))
            lines.append("| " + " | ".join(padded[:len(header)]) + " |")

        lines.append("")

    wb.close()

    content = "\n".join(lines)
    if len(content) > 8000:
        content = content[:8000] + "\n\n[... Data truncated ...]"

    return {
        "success": True,
        "filename": filename,
        "filetype": os.path.splitext(filename)[1].lower(),
        "content": content,
        "error": None,
    }


# ── Plain Text Parser ────────────────────────────────────────────
def _parse_text(file_storage, filename):
    raw = file_storage.stream.read()
    for encoding in ("utf-8", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw.decode("utf-8", errors="replace")

    if len(text) > 8000:
        text = text[:8000] + "\n\n[... Text truncated ...]"

    return {
        "success": True,
        "filename": filename,
        "filetype": ".txt",
        "content": f"[Text File: {filename}]\n\n{text}",
        "error": None,
    }


# ── JSON Parser ──────────────────────────────────────────────────
def _parse_json(file_storage, filename):
    raw = file_storage.stream.read()
    try:
        data = json.loads(raw.decode("utf-8"))
        formatted = json.dumps(data, indent=2)
        if len(formatted) > 8000:
            formatted = formatted[:8000] + "\n\n[... Data truncated ...]"
        return {
            "success": True,
            "filename": filename,
            "filetype": ".json",
            "content": f"[JSON File: {filename}]\n\n{formatted}",
            "error": None,
        }
    except json.JSONDecodeError as e:
        return {
            "success": False,
            "filename": filename,
            "filetype": ".json",
            "content": "",
            "error": f"Invalid JSON: {str(e)}",
        }
